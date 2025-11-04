import csv
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

#!/usr/bin/env python3

def create_excel_report(csv_file):
    """Generate an Excel report from a CSV audit file."""
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Rapport Audit ANSSI"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    failure_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Read CSV and populate worksheet
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f, delimiter=';')
        
        # Headers
        headers = ['ID', 'Nom', 'Description', 'Niveau', 'Succès', 'Commentaire']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Data rows
        for row_idx, row in enumerate(reader, 2):
            ws.cell(row=row_idx, column=1, value=row['id']).border = border
            ws.cell(row=row_idx, column=2, value=row['nom']).border = border
            ws.cell(row=row_idx, column=3, value=row['description']).border = border
            ws.cell(row=row_idx, column=4, value=int(row['niveau'])).border = border
            
            # Success column with color
            success_cell = ws.cell(row=row_idx, column=5, value=row['succes'])
            success_cell.border = border
            success_cell.alignment = Alignment(horizontal='center')
            if row['succes'].lower() == 'true':
                success_cell.fill = success_fill
            else:
                success_cell.fill = failure_fill
            
            comment_cell = ws.cell(row=row_idx, column=6, value=row['commentaire'])
            comment_cell.border = border
            comment_cell.alignment = Alignment(wrap_text=True)
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 70
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 60
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save workbook
    output_file = "rapports/" + csv_file.rsplit('.', 1)[0] + '_rapport.xlsx'
    os.makedirs("rapports", exist_ok=True)
    wb.save(output_file)
    print(f"Rapport généré : {output_file}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python rapport-xslx.py <fichier.csv>")
        sys.exit(1)
    
    create_excel_report(sys.argv[1])